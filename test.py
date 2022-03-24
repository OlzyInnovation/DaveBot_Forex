from os.path import exists as file_exists
import openpyxl
from openpyxl import load_workbook


symbols: list[str] = ['BINANCE_SPOT_MATIC_USDT','BINANCE_SPOT_LUNA_USDT', 'BINANCE_SPOT_ETH_USDT', 'BINANCE_SPOT_ADA_USDT', 'BINANCE_SPOT_SOL_USDT', 'BINANCE_SPOT_CRV_USDT','BINANCE_SPOT_BTC_USDT' ]

periods: list[str] = ['5MIN', '15MIN', '1HRS', '4HRS', '1DAY']


def split(word):
    return [char for char in word]


def get_coordinates(symbol, period, ):
    pass


def workbook() -> None:
    if not file_exists('demo.xlsx'):         
        # Call Workbook() function of openpyxl 
        # to create a new blank Workbook object
            wb = openpyxl.Workbook()

            # Save the blank workbook
            wb.save('demo.xlsx')
    wb = load_workbook('demo.xlsx')
    work_sheet = wb.active

    # Append Symbols to work_sheet
    for i in range(0, len(symbols)):
        symbol_to_append = symbols[i].split('_')
        work_sheet[f'A{i + 2}'].value = symbol_to_append[-2] + '/' + symbol_to_append[-1]
    
    # Append Periods to work_sheet
    for i in range(0, len(periods)):            
        work_sheet.cell(row=1, column=i + 2).value = periods[i]

    for row in work_sheet.iter_rows():
        for cell in row:
            if cell.value == 'MATIC/USDT':
                order = split(cell.coordinate)
                print('{} is order'.format(order))
                print(chr(ord(order[0]) + 1) + order[1])
                work_sheet[chr(ord(order[0]) + 1) + order[1]] = '↘'
            if cell.value == 'LUNA/USDT':
                order = split(cell.coordinate)
                print('{} is order'.format(order))
                print(chr(ord(order[0]) + 1) + order[1])
                work_sheet[chr(ord(order[0]) + 2) + order[1]] = '↗'
            if cell.value == 'BTC/USDT':
                order = split(cell.coordinate)
                print('{} is order'.format(order))
                print(chr(ord(order[0]) + 1) + order[1])
                work_sheet[chr(ord(order[0]) + 5) + order[1]] = '↘'
    wb.save('demo.xlsx')


workbook()
