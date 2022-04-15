from os.path import exists as file_exists
import openpyxl
from openpyxl import load_workbook

from symbols import symbols
from periods import periods


def split(word):
    return [char for char in word]


def workbook(symbol: str, period: str, price: float, ma: float, success: float, unsuccess: float, blackX: float):
    if not file_exists('demo.xlsx'):         
        # Call Workbook() function of openpyxl 
        # to create a new blank Workbook object
        wb = openpyxl.Workbook()
        work_sheet = wb.active
        # Append Symbols to work_sheet
        for i in range(0, len(symbols)):
            
            work_sheet[f'A{i + 2}'].value = symbols[i]
        
        # Append Periods to work_sheet
        for i in range(0, len(periods)):            
            work_sheet.cell(row=1, column=i + 2).value = periods[i]

        # Save the blank workbook
        wb.save('demo.xlsx')

    wb = load_workbook('demo.xlsx')
    work_sheet = wb.active

    for i in range(0, len(symbols)):
        work_sheet[f'A{i + 2}'].value = symbols[i]
                
    # Append Periods to work_sheet
    for i in range(0, len(periods)):            
        work_sheet.cell(row=1, column=i + 2).value = periods[i]

    for row in work_sheet.iter_rows():
        difference = price > ma
        message = ''
        up = '↗'
        down = '↘'
        for cell in row:
            if cell.value == symbol:
                order = split(cell.coordinate)
                if period == '5MIN':
                    if difference:
                        message = f'{success}/{blackX}    {up}'
                        work_sheet[chr(ord(order[0]) + 1) + order[1]] = message
                    else:
                        message = f'{success}/{blackX}    {down}'
                        work_sheet[chr(ord(order[0]) + 1) + order[1]] = message
                elif period == '15MIN':
                    if difference:
                        message = f'{success}/{blackX}    {up}'
                        work_sheet[chr(ord(order[0]) + 2) + order[1]] = message
                    else:
                        message = f'{success}/{blackX}    {down}'
                        work_sheet[chr(ord(order[0]) + 2) + order[1]] = message
                elif period == '1HRS':
                    if difference:
                        message = f'{success}/{blackX}    {up}'
                        work_sheet[chr(ord(order[0]) + 3) + order[1]] = message
                    else:
                        message = f'{success}/{blackX}    {down}'
                        work_sheet[chr(ord(order[0]) + 3) + order[1]] = message
                elif period == '4HRS':
                    if difference:
                        message = f'{success}/{blackX}    {up}'
                        work_sheet[chr(ord(order[0]) + 4) + order[1]] = message
                    else:
                        message = f'{success}/{blackX}    {down}'
                        work_sheet[chr(ord(order[0]) + 4) + order[1]] = message
                elif period == '1DAY':
                    if difference:
                        message = f'{success}/{blackX}    {up}'
                        work_sheet[chr(ord(order[0]) + 5) + order[1]] = message
                    else:
                        message = f'{success}/{blackX}    {down}'
                        work_sheet[chr(ord(order[0]) + 5) + order[1]] = message

    wb.save('demo.xlsx')
